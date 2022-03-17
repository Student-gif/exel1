import xlsxwriter
import openpyxl
Spo = 'SPO'
plan = 'plan'

def parseExelDic(SheetName):
    wb = openpyxl.load_workbook('exelDb.xlsx')
    sheet = wb[SheetName]
    key_and_names = {}

    for i in range(1,63):
        g = 2
        while sheet.cell(row=g, column=i).value != None:
            key = sheet.cell(row=1, column=i).value
            if sheet.cell(row=g, column=i) != None:
                lessons = sheet.cell(row=g, column=i).value
      
            g+=1

            if not key in key_and_names:
                key_and_names[key]=[]
            key_and_names[key].append(lessons)   
    return key_and_names
    

SpoDic = parseExelDic(Spo)
plan = parseExelDic(plan)
